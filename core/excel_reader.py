"""Leitura genérica das planilhas .xlsx, isolando as abas de dados reais"""
from __future__ import annotations

from typing import Dict, List, Optional

import openpyxl
import pandas as pd

from core.utils import normalize

# Algumas abas escrevem o mesmo cabeçalho com palavras diferentes (ex: as
# abas JUN26/JUL26 da planilha de laudos usam "ENTRADA DO LAUDO" em vez de
# "ENTRADA DE LAUDO" usado nas demais). Sem isso, essas abas viram colunas
# separadas depois de juntar tudo, e as linhas ficam com o status "vazio" —
# sendo descartadas nos filtros. Adicione aqui outras variações que
# aparecerem no futuro (a comparação já ignora acento/caixa/espaço).
HEADER_ALIASES = {
    normalize("ENTRADA DO LAUDO"): normalize("ENTRADA DE LAUDO"),
}


def _canonical_header(texto: str) -> str:
    """Normaliza um cabeçalho (acento/caixa/espaço) e aplica os apelidos
    conhecidos, para que a mesma coluna vinda de abas diferentes sempre
    caia na mesma coluna do DataFrame final."""
    chave = normalize(texto)
    return HEADER_ALIASES.get(chave, chave)


def _find_header_row(rows: List[tuple], required_headers: List[str]) -> Optional[int]:
    """Procura, nas primeiras linhas de uma aba, aquela que contém todos os
    cabeçalhos exigidos (comparação exata, sem normalizar, pois os
    cabeçalhos das planilhas reais já vêm em maiúsculas)."""
    required_canonicos = [_canonical_header(h) for h in required_headers]
    for i, row in enumerate(rows[:5]):
        values = [_canonical_header(c) for c in row if c is not None]
        # Comparação EXATA (não substring): evita que abas de controle/dashboard
        # com colunas parecidas (ex: "TIPOS DE LAUDO", "TIPO DE LAUDO (canônicos)")
        # sejam confundidas com abas de dados reais.
        if all(h in values for h in required_canonicos):
            return i
    return None


def load_data_sheets(path: str, required_headers: List[str]) -> pd.DataFrame:
    """Lê todas as abas do arquivo que contenham as colunas exigidas
    (ex: 'EMPRESA' e 'TIPO DE LAUDO') e devolve tudo empilhado num único
    DataFrame, com uma coluna extra '_ABA' indicando de qual mês veio.

    Abas de controle/dashboard (que não têm essas colunas) são ignoradas
    automaticamente — não é preciso listar seus nomes.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    frames = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header_idx = _find_header_row(rows, required_headers)
        if header_idx is None:
            continue
        header = [
            _canonical_header(c) if c is not None else f"col_{i}"
            for i, c in enumerate(rows[header_idx])
        ]
        # nomes de coluna duplicados (ex: duas colunas "DATA") viram DATA, DATA_2
        seen: Dict[str, int] = {}
        cols = []
        for h in header:
            seen[h] = seen.get(h, 0) + 1
            cols.append(h if seen[h] == 1 else f"{h}_{seen[h]}")
        data_rows = rows[header_idx + 1:]
        df = pd.DataFrame(data_rows, columns=cols)
        df["_ABA"] = sheet_name
        # descarta linhas totalmente vazias
        df = df.dropna(how="all", subset=[c for c in cols if c != "_ABA"])
        frames.append(df)
    if not frames:
        return pd.DataFrame()
    resultado = pd.concat(frames, ignore_index=True)
    # Algumas planilhas mantêm uma aba "mestre" (ex: AGENDAMENTO) além das
    # abas por mês, e o mesmo registro acaba em duas abas ao mesmo tempo —
    # o que duplicaria a contagem no relatório. Remove linhas idênticas
    # (ignorando só a coluna _ABA), mantendo a primeira ocorrência.
    colunas_para_comparar = [c for c in resultado.columns if c != "_ABA"]
    resultado = resultado.drop_duplicates(subset=colunas_para_comparar, keep="first")
    return resultado.reset_index(drop=True)


def list_sheet_names(path: str) -> List[str]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    return wb.sheetnames
